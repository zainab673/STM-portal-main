from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from .models import UserProfile, Teacher, StudentProfile  # Import StudentProfile
from django.http import HttpResponseRedirect, Http404
from django.urls import reverse
from .models import Timetable, ContactMessage, Notification
from openpyxl import Workbook
from django.http import HttpResponse, FileResponse, JsonResponse
import os
from django.conf import settings
import platform
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
import json
import requests
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, F
from django.db.models.functions import Length
from .forms import TeacherForm, StudentProfileForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
import logging
from django.db.models import Case, When, Value, IntegerField

logger = logging.getLogger(__name__)

# Hardcoded verification codes for each role
VERIFICATION_CODES = {
    "student": "STUDENT123",
    "teacher": "TEACHER456",
    "admin": "ADMIN789",
}


def index(request):
    # Check if the role is passed in the URL
    role = request.GET.get('role', None)
    if role:
        # Store the role in the session
        request.session['role'] = role
        print(f"Role stored in session: {role}")
        return redirect('signup')  # Redirect to signup page
    return render(request, 'index.html')


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        # Get role from POST or fallback to GET
        selected_role = request.POST.get('role') or request.GET.get('role') or ''
        selected_role = selected_role.lower().strip()

        print(f"🔍 Role from POST: {request.POST.get('role')}")
        print(f"🔍 Role from GET: {request.GET.get('role')}")
        print(f"✅ Final selected_role: {selected_role}")

        try:
            user_obj = UserProfile.objects.get(username=username)
        except UserProfile.DoesNotExist:
            return render(request, 'login.html', {
                'error': 'User not found. Please create an account.',
                'role': selected_role
            })

        user = authenticate(request, username=username, password=password)

        if user is not None:
            print(f"✅ User role from DB: {user.role.lower()}")
            if user.role.lower().strip() != selected_role:
                return render(request, 'login.html', {
                    'error': 'Role mismatch. Please select the correct role.',
                    'role': selected_role
                })

            login(request, user)
            request.session['role'] = user.role
            print(f"🔓 Logged in as: {user.username}, Role: {user.role}")

            # Redirect by role
            if user.role.lower() == 'student':
                return redirect('student_panel')
            elif user.role.lower() == 'teacher':
                return redirect('teacher_panel')
            elif user.role.lower() == 'admin':
                return redirect('admin_panel')
        else:
            return render(request, 'login.html', {
                'error': 'Incorrect password. Please try again.',
                'role': selected_role
            })

    # Handle GET request
    role = request.GET.get('role', '').lower()
    return render(request, 'login.html', {'role': role})


def signup_view(request):
    print(f"🔍 Session Data: {dict(request.session)}")
    role = request.session.get('role', None)

    if not role:
        role = request.POST.get('role', '').lower()  # Fetch role from POST if not in session
        if role:
            request.session['role'] = role  # Store the role in session for future requests
        else:
            print("No role found! Redirecting to index...")
            return redirect('index')

    print(f"Fetched role from session or POST: {role}")

    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        verification_code = request.POST['verification_code']

        expected_code = VERIFICATION_CODES.get(role, '')
        print(f"Expected Code: {expected_code}, Received: {verification_code}")

        # Check if the verification code is correct
        if verification_code != expected_code:
            print("Invalid verification code!")
            messages.error(request, 'Invalid verification code! Please enter the correct code.')
            return render(request, 'signup.html', {'role': role})

        if UserProfile.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists. Please choose another one.')
            return render(request, 'signup.html', {'role': role})

        # Create user and assign role
        try:
            user = UserProfile.objects.create_user(username=username, email=email, password=password)
            user.role = role
            user.save()
            print(f"User {username} created successfully!")

            if role == 'teacher':
                Teacher.objects.create(user=user, verification_code=verification_code)
                print(f"Teacher profile created for user: {username}")
            elif role == 'student':
                StudentProfile.objects.create(user=user, semester=1)
                print(f"Student profile created for user: {username}")
        except Exception as e:
            messages.error(request, f'An error occurred during signup: {e}')
            return render(request, 'signup.html', {'role': role})

        # After successful user creation and profile creation
        login(request, user)
        print(f"🔍 User {username} logged in with ID: {request.session.get('_auth_user_id')}")
        print(f"🔍 Role assigned: {user.role}")
        messages.success(request, '🎉 Account created successfully!')

        # Check the role and redirect to the correct dashboard
        if user.role == 'student':
            print("Redirecting to student dashboard.")
            return redirect('student_panel')
        elif user.role == 'teacher':
            return redirect('teacher_panel')
        elif user.role == 'admin':
            return redirect('admin_panel')
    return render(request, 'signup.html', {'role': role})


def student_panel(request):
    return render(request, "studentpanel.html")


def teacher_panel(request):
    return render(request, "teacherpanel.html")


def admin_panel(request):
    unread_messages = ContactMessage.objects.filter(is_read=False).order_by('-timestamp')
    unread_messages_count = unread_messages.count()
    context = {
        'unread_messages': unread_messages,
        'unread_messages_count': unread_messages_count,
    }
    return render(request, "adminpanel.html", context)


def select_semester(request):
    return render(request, 'selectsem.html')


# EXCEL SHEET FOR ATTENDANCE
def open_attendance(request):
    if request.method == "POST":
        semester = request.POST.get("semester")
        subject = request.POST.get("subject")

        if not semester or not subject:
            return HttpResponse("Please select both semester and subject.")

        # File name and folder
        file_name = f"{subject.replace(' ', '_')}_Semester{semester}.xlsx"
        folder_path = os.path.join(settings.MEDIA_ROOT, "attendance_files")
        file_path = os.path.join(folder_path, file_name)

        # Ensure folder exists
        os.makedirs(folder_path, exist_ok=True)

        # If file doesn't exist, create it
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"

            # Write only Semester and Subject
            ws['A1'] = f"Semester: {semester}"
            ws['A2'] = f"Subject: {subject}"

            wb.save(file_path)

        # If the teacher wants to download the file
        if 'download' in request.POST:
            if os.path.exists(file_path):
                try:
                    # Serve the file using FileResponse
                    response = FileResponse(open(file_path, 'rb'),
                                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
                    return response
                except Exception as e:
                    return HttpResponse(f"Error occurred while trying to serve the file: {str(e)}")
            else:
                return HttpResponse("File does not exist. Please try again later.")

        # If the teacher wants to open the file (for Windows only)
        if platform.system() == "Windows":
            os.startfile(file_path)
            return HttpResponse("Excel file opened successfully!")

        # For non-Windows systems
        return HttpResponse("File created. Please open it manually (non-Windows system).")

    return HttpResponse("Invalid request method.")


def chatbot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            message = data.get("message", "")

            api_key = "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"  # Replace with your API key

            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:8000/",
                # If you are testing locally
                "X-Title": "Scholar Sphere Chatbot"  # Your project title
            }

            body = {
                "model": "openai/gpt-3.5-turbo",  # 📢 Make sure model name is correct
                "messages": [
                    {"role": "system",
                     "content": "You are an educational assistant helping students and staff."},
                    {"role": "user", "content": message}
                ]
            }

            response = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers=headers,
                json=body
            )

            # Log the response for debugging
            result = response.json()

            if response.status_code == 200:
                answer = result['choices'][0]['message']['content'].strip()
                return JsonResponse({'response': answer})
            else:
                return JsonResponse({'error': 'API error', 'details': result}, status=500)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


def submit_contact_form(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')

        if name and email and message:
            ContactMessage.objects.create(name=name, email=email, message=message)
            return HttpResponse("success")
        else:
            return HttpResponse("error", status=400)
    else:
        return redirect('index')


def read_message(request, message_id):
    message = get_object_or_404(ContactMessage, id=message_id)
    message.is_read = True
    message.save()
    return redirect('admin_panel')


@csrf_protect
def delete_message(request, message_id):
    if request.method == 'POST':
        message = get_object_or_404(ContactMessage, id=message_id)
        message.delete()
    return redirect('admin_panel')


def teacher_list(request):
    teachers = Teacher.objects.all()

    # Search
    search_term = request.GET.get('search')
    if search_term:
        teachers = teachers.filter(
            Q(user__username__icontains=search_term) |
            Q(user__email__icontains=search_term) |
            Q(subject_taught__icontains=search_term) |
            Q(department__icontains=search_term)
        )
        # Annotate with a 'relevance' score. Prioritize username matches.
        teachers = teachers.annotate(
            relevance=Case(
                When(user__username__icontains=search_term, then=Value(10)),
                # High priority
                When(user__email__icontains=search_term, then=Value(5)),
                # Medium priority
                default=Value(1),
                output_field=IntegerField(),
            ),
            name_length=Length('user__username')
        ).order_by('-relevance',
                     'name_length')  # Important: Order by relevance and then name_length
    else:
        teachers = teachers.annotate(
            relevance=Value(0, output_field=IntegerField()),
            name_length=Length('user__username')
        )

    # Sorting
    sort_by = request.GET.get('sort')
    if sort_by:
        if sort_by == 'username':
            teachers = teachers.order_by('user__username')
        elif sort_by == 'email':
            teachers = teachers.order_by('user__email')
        elif sort_by == 'department':
            teachers = teachers.order_by('department')
        elif sort_by == 'cnic':
            teachers = teachers.order_by('cnic')
        elif sort_by == 'contact_number':
            teachers = teachers.order_by('contact_number')
        elif sort_by == 'subject_taught':
            teachers = teachers.order_by('subject_taught')
        else:
            teachers = teachers.order_by('pk')  # Default to primary key if sort is invalid
    else:
        teachers = teachers.order_by('pk')  # Default ordering

    # Pagination
    paginator = Paginator(teachers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_by': sort_by,
    }
    return render(request, 'admin_teachers.html', context)


def teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('teacher_list')
    else:
        form = TeacherForm(instance=teacher)
    return render(request, 'admin_teacher_form.html', {'form': form, 'teacher': teacher})


def teacher_delete(request, pk):
    try:
        teacher = get_object_or_404(Teacher, pk=pk)
        teacher_username = "this teacher"  # Default username if no related user
        if teacher.user:
            teacher_username = teacher.user.username

        if request.method == 'GET':
            teacher.delete()
            messages.success(request, f"Teacher '{teacher_username}' deleted successfully.")
            return redirect(reverse('teacher_list'))
        return redirect(reverse('teacher_list'))
    except Http404:
        logger.error(f"Teacher with pk={pk} not found for deletion.")
        messages.error(request, "Deleted Successfully.")
        return redirect(reverse('teacher_list'))


@login_required
def teacher_dashboard(request):
    context = {
        'is_admin': False,
    }
    return render(request, 'teacher_dashboard.html', context)


@login_required
def student_dashboard(request):
    context = {
        'is_admin': False,
    }
    return render(request, 'student_dashboard.html', context)


@login_required
def admin_dashboard(request):
    is_admin = False
    if request.user.is_superuser or request.user.groups.filter(
            name='Admin').exists():
        is_admin = True
    context = {
        'is_admin': is_admin,
    }
    return render(request, 'admin_dashboard.html', context)


@login_required
def admin_timetable(request):
    is_admin = False
    if request.user.is_superuser or request.user.groups.filter(
            name='Admin').exists():
        is_admin = True
    context = {
        'is_admin': is_admin,
        'timetable': Timetable.objects.all()  # Or filter it as needed
    }
    print(f"Admin Timetable Context: {context}")  # Add this line
    return render(request, 'timetable.html', context)


@login_required
def teacher_timetable(request):
    context = {
        'is_admin': False,
        'timetable': Timetable.objects.all()  # Or filter for the teacher's classes
    }
    return render(request, 'timetable.html', context)


@login_required
def student_timetable(request):
    context = {
        'is_admin': False,
        'timetable': Timetable.objects.all()  # Or filter for student's classes
    }
    return render(request, 'timetable.html', context)
# student_management_app/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Timetable


@login_required
def admin_timetable(request):
    context = {
        'timetable': Timetable.objects.all()
    }
    return render(request, 'timetable.html', context)


@login_required
def teacher_timetable(request):
    context = {
        'timetable': Timetable.objects.all()  # Or filter as needed
    }
    return render(request, 'timetable.html', context)


@login_required
def student_timetable(request):
    context = {
        'timetable': Timetable.objects.all()  # Or filter as needed
    }
    return render(request, 'timetable.html', context)
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Timetable, Teacher, Notification
from django.http import JsonResponse


# ... your existing admin timetable view ...

@user_passes_test(lambda u: u.is_superuser)  # Example: Only admin can save timetable
def save_timetable_view(request):
    if request.method == 'POST':
        # ... your logic to save the timetable ...
        # After saving, identify affected teachers and create notifications
        affected_teachers = Teacher.objects.filter(
            subject_taught__in=[])  # Example: Filter teachers by taught subjects
        semester = request.POST.get('semester')  # Example: Get the affected semester

        for teacher in affected_teachers:
            Notification.objects.create(
                user=teacher.user,
                message=f"Timetable updated for Semester {semester}.",
                related_semester=int(semester) if semester else None
            )
        # Optionally send a real-time notification using Django Channels

        return JsonResponse(
            {'success': True, 'message': 'Timetable saved and notifications sent.'})
    # ... handle GET request ...


@login_required
def get_notifications(request):
    unread_notifications = Notification.objects.filter(
        user=request.user, is_read=False)
    notifications_data = [{
        'message': n.message,
        'timestamp': n.timestamp.strftime('%Y-%m-%d %H:%M'),
        'related_semester': n.related_semester
    } for n in unread_notifications]
    return JsonResponse({
        'unread_count': unread_notifications.count(),
        'notifications': notifications_data
    })


@login_required
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user,
                                    is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from .models import StudentProfile
from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from .models import StudentProfile


def student_list(request):  # Make sure this function exists and is named correctly
    students = StudentProfile.objects.all()
    search_term = request.GET.get('search')
    sort_by = request.GET.get('sort')

    if search_term:
        students = students.filter(
            Q(user__username__icontains=search_term) |
            Q(user__email__icontains=search_term) |
            Q(cnic__icontains=search_term) |
            Q(contact_number__icontains=search_term)
        )

    if sort_by:
        if sort_by == 'user__username':
            students = students.order_by('user__username')
        elif sort_by == 'user__email':
            students = students.order_by('user__email')
        elif sort_by == 'semester':
            students = students.order_by('semester')
        elif sort_by == 'contact_number':
            students = students.order_by('contact_number')
        elif sort_by == 'cnic':
            students = students.order_by('cnic')
        else:
            students = students.order_by('pk')  # default sort
    else:
        students = students.order_by('pk')

    paginator = Paginator(students, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'search_term': search_term,
        'sort_by': sort_by,
    }
    return render(request, 'admin_students.html', context)


from django.shortcuts import render, get_object_or_404
from .forms import StudentProfileForm
from .models import UserProfile, StudentProfile


def student_edit(request, pk):
    user = get_object_or_404(UserProfile, pk=pk)
    student_profile = StudentProfile.objects.get(user=user)
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, instance=student_profile)
        if form.is_valid():
            form.save()
            #  messages.success(request, f"{user.username}'s profile updated successfully!") # Removed import
            return redirect('student_list')
    else:
        form = StudentProfileForm(instance=student_profile)
    return render(request, 'admin_students_form.html',
                  {'form': form, 'student': user})


@login_required
@user_passes_test(lambda u: u.role == 'admin')
def student_delete(request, pk):
    try:
        user_to_delete = get_object_or_404(UserProfile, pk=pk)
        username = user_to_delete.username
        user_to_delete.delete()
        messages.success(request, f"Student '{username}' deleted successfully!")
        return redirect('student_list')
    except Http404:
        messages.error(request, "Student not found!")
        return redirect('student_list')
# student_management_app/views.py

from django.shortcuts import render
# You might have other imports here, e.g., for models, forms, etc.
# from .models import Teacher, Course, Assignment # Example if you use Django ORM

def admin_courses(request):
    """
    Renders the page for managing teacher courses.
    This view will serve the HTML content that uses Firestore for data management.
    """
    # In a real Django application, you might fetch initial data from your Django models
    # and pass it to the template context if needed.
    # For this Firestore-based page, most data loading happens on the client-side.

    # Example of passing context (optional, adjust as per your Django setup)
    context = {
        'page_title': 'Manage Teacher Courses',
        # 'teachers_data': Teacher.objects.all(), # Example if fetching from Django DB
    }
    return render(request, 'Manage_Teacher_Courses.html', context)

# Add other view functions you might have here, e.g.:
# def student_list(request):
#     return render(request, 'student_list.html')

# def teacher_list(request):
#     return render(request, 'teacher_list.html')

# def admin_timetable(request):
#     return render(request, 'admin_timetable.html')

# def delete_message(request, message_id):
#     # Implement your message deletion logic here
#     # This would typically involve interacting with your Django models
#     # For now, a placeholder response:
#     from django.http import JsonResponse
#     if request.method == 'POST':
#         # Simulate deletion success
#         return JsonResponse({'status': 'success', 'message': f'Message {message_id} deleted.'})
#     return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)


